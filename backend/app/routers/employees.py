from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session

from .. import models, schemas
from ..auth import generate_username, hash_password, require_admin, get_current_user
from ..config import AREAS
from ..database import get_db

router = APIRouter(prefix="/api/employees", tags=["employees"])

PASSWORD_INICIAL = "1234"


def _employee_out(emp: models.Employee, db: Session) -> schemas.EmployeeOut:
    pendientes = (
        db.query(models.VacationRequest)
        .filter(
            models.VacationRequest.employee_id == emp.id,
            models.VacationRequest.estado == models.RequestStatus.pendiente,
        )
        .count()
    )
    return schemas.EmployeeOut(
        id=emp.id,
        nombre=emp.nombre,
        apellido=emp.apellido,
        area=emp.area,
        fecha_ingreso=emp.fecha_ingreso,
        dias_usados=emp.dias_usados,
        tipo_contrato=emp.tipo_contrato,
        activo=emp.activo,
        username=emp.user.username if emp.user else None,
        meses_trabajados=emp.meses_trabajados(),
        dias_acumulados=emp.dias_acumulados(),
        saldo=emp.saldo(),
        solicitudes_pendientes=pendientes,
    )


def _create_employee_with_user(
    db: Session, data: schemas.EmployeeCreate
) -> tuple[models.Employee, str]:
    emp = models.Employee(
        nombre=data.nombre.strip(),
        apellido=data.apellido.strip(),
        area=data.area.strip(),
        fecha_ingreso=data.fecha_ingreso,
        dias_usados=data.dias_usados,
        tipo_contrato=data.tipo_contrato,
    )
    db.add(emp)
    db.flush()
    username = generate_username(db, emp.nombre, emp.apellido)
    user = models.User(
        username=username,
        password_hash=hash_password(PASSWORD_INICIAL),
        role=models.Role.empleado,
        employee_id=emp.id,
    )
    db.add(user)
    return emp, username


@router.get("/areas")
def list_areas():
    return AREAS


@router.get("", response_model=list[schemas.EmployeeOut])
def list_employees(
    q: str | None = None,
    area: str | None = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    query = db.query(models.Employee).filter(models.Employee.activo == True)  # noqa: E712
    if area:
        query = query.filter(models.Employee.area == area)
    employees = query.order_by(models.Employee.apellido).all()
    if q:
        ql = q.lower()
        employees = [
            e for e in employees
            if ql in e.nombre.lower() or ql in e.apellido.lower()
        ]
    return [_employee_out(e, db) for e in employees]


@router.post("", response_model=schemas.EmployeeCreated, status_code=201)
def create_employee(
    data: schemas.EmployeeCreate,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    emp, username = _create_employee_with_user(db, data)
    db.commit()
    db.refresh(emp)
    return schemas.EmployeeCreated(
        employee=_employee_out(emp, db),
        username=username,
        password_inicial=PASSWORD_INICIAL,
    )


@router.put("/{employee_id}", response_model=schemas.EmployeeOut)
def update_employee(
    employee_id: int,
    data: schemas.EmployeeUpdate,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    emp = db.get(models.Employee, employee_id)
    if emp is None:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(emp, field, value)
    db.commit()
    db.refresh(emp)
    return _employee_out(emp, db)


@router.delete("/{employee_id}")
def deactivate_employee(
    employee_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    emp = db.get(models.Employee, employee_id)
    if emp is None:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    emp.activo = False
    db.commit()
    return {"ok": True}


@router.post("/import", response_model=schemas.ImportResult)
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    """Importa empleados desde .xlsx con columnas: Nombre, Área, FechaIngreso, DíasUsados."""
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx")

    import io
    from openpyxl import load_workbook

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel")

    from ..auth import slugify_name

    ws = wb.active
    # slugify: sin tildes, espacios ni símbolos, para tolerar variantes de encabezado
    headers = [slugify_name(str(c.value)) if c.value else "" for c in ws[1]]

    def col(*names):
        for name in names:
            if name in headers:
                return headers.index(name)
        return None

    idx_nombre = col("nombre")
    idx_area = col("area", "rea")  # "rea" cubre archivos con la Á corrupta
    idx_fecha = col("fechaingreso")
    idx_dias = col("diasusados", "dasusados")

    if idx_nombre is None or idx_area is None or idx_fecha is None:
        raise HTTPException(
            status_code=400,
            detail="Columnas requeridas: Nombre, Área, FechaIngreso, DíasUsados",
        )

    importados = 0
    errores: list[str] = []
    credenciales: list[dict] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None for v in row):
            continue
        try:
            nombre_completo = str(row[idx_nombre]).strip()
            partes = nombre_completo.split()
            if len(partes) >= 2:
                nombre, apellido = " ".join(partes[:-1]), partes[-1]
            else:
                nombre, apellido = nombre_completo, ""

            area = str(row[idx_area]).strip()
            raw_fecha = row[idx_fecha]
            if isinstance(raw_fecha, datetime):
                fecha = raw_fecha.date()
            else:
                fecha = datetime.strptime(str(raw_fecha).strip()[:10], "%Y-%m-%d").date()

            dias_usados = float(row[idx_dias] or 0) if idx_dias is not None else 0.0

            emp, username = _create_employee_with_user(
                db,
                schemas.EmployeeCreate(
                    nombre=nombre,
                    apellido=apellido,
                    area=area,
                    fecha_ingreso=fecha,
                    dias_usados=dias_usados,
                ),
            )
            credenciales.append(
                {"empleado": emp.nombre_completo, "usuario": username, "contraseña": PASSWORD_INICIAL}
            )
            importados += 1
        except Exception as exc:
            errores.append(f"Fila {i}: {exc}")

    db.commit()
    return schemas.ImportResult(importados=importados, errores=errores, credenciales=credenciales)


@router.get("/me", response_model=schemas.EmployeeDashboard)
def my_dashboard(
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.employee is None:
        raise HTTPException(status_code=404, detail="El usuario no tiene ficha de empleado")
    emp = user.employee
    return schemas.EmployeeDashboard(
        nombre_completo=emp.nombre_completo,
        area=emp.area,
        fecha_ingreso=emp.fecha_ingreso,
        meses_trabajados=emp.meses_trabajados(),
        dias_acumulados=emp.dias_acumulados(),
        dias_usados=emp.dias_usados,
        saldo=emp.saldo(),
    )
